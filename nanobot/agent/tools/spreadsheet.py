"""Spreadsheet tools: read and write CSV/XLSX files."""

import csv
import io
import json
from pathlib import Path
from typing import Any

from nanobot.agent.tools.base import Tool


def _resolve_path(path: str, allowed_dir: Path | None = None) -> Path:
    """Resolve path and optionally enforce directory restriction."""
    resolved = Path(path).expanduser().resolve()
    if allowed_dir and not str(resolved).startswith(str(allowed_dir.resolve())):
        raise PermissionError(f"Path {path} is outside allowed directory {allowed_dir}")
    return resolved


class SpreadsheetReadTool(Tool):
    """Read data from CSV or XLSX spreadsheet files."""

    def __init__(self, allowed_dir: Path | None = None):
        self._allowed_dir = allowed_dir

    @property
    def name(self) -> str:
        return "spreadsheet_read"

    @property
    def description(self) -> str:
        return (
            "Read data from a spreadsheet file (CSV or XLSX). "
            "Returns data as a formatted table. Optionally specify sheet name and cell range."
        )

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Path to the spreadsheet file (.csv or .xlsx)"},
                "sheet": {"type": "string", "description": "Sheet name (for XLSX files, default: first sheet)"},
                "range": {"type": "string", "description": "Cell range, e.g., 'A1:D10' (optional)"},
            },
            "required": ["path"],
        }

    async def execute(self, **kwargs: Any) -> str:
        path_str = kwargs.get("path", "").strip()
        sheet_name = kwargs.get("sheet", "").strip() or None
        cell_range = kwargs.get("range", "").strip() or None

        if not path_str:
            return "Error: No file path provided."

        try:
            file_path = _resolve_path(path_str, self._allowed_dir)
        except PermissionError as e:
            return f"Error: {e}"

        if not file_path.exists():
            return f"Error: File not found: {path_str}"

        suffix = file_path.suffix.lower()

        try:
            if suffix == ".csv":
                return self._read_csv(file_path, cell_range)
            elif suffix in (".xlsx", ".xls"):
                return self._read_xlsx(file_path, sheet_name, cell_range)
            else:
                return f"Error: Unsupported file type '{suffix}'. Use .csv or .xlsx."
        except Exception as e:
            return f"Error reading spreadsheet: {e}"

    def _read_csv(self, path: Path, cell_range: str | None) -> str:
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            rows = list(reader)

        if cell_range:
            rows = self._apply_range(rows, cell_range)

        return self._format_table(rows, f"CSV: {path.name}")

    def _read_xlsx(self, path: Path, sheet_name: str | None, cell_range: str | None) -> str:
        try:
            import openpyxl
        except ImportError:
            return "Error: openpyxl is required for XLSX files. Install with: pip install openpyxl"

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                return f"Error: Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}"
            ws = wb[sheet_name]
        else:
            ws = wb.active
            sheet_name = ws.title

        if cell_range:
            rows = []
            for row in ws[cell_range]:
                rows.append([str(cell.value) if cell.value is not None else "" for cell in row])
        else:
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(v) if v is not None else "" for v in row])

        wb.close()
        return self._format_table(rows, f"XLSX: {path.name} [{sheet_name}]")

    @staticmethod
    def _apply_range(rows: list[list[str]], cell_range: str) -> list[list[str]]:
        """Apply a simple A1:D10 style range to row data."""
        import re
        match = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', cell_range.upper())
        if not match:
            return rows

        def col_to_idx(col: str) -> int:
            idx = 0
            for c in col:
                idx = idx * 26 + (ord(c) - ord('A'))
            return idx

        c1, r1, c2, r2 = match.groups()
        row_start = int(r1) - 1
        row_end = int(r2)
        col_start = col_to_idx(c1)
        col_end = col_to_idx(c2) + 1

        return [row[col_start:col_end] for row in rows[row_start:row_end]]

    @staticmethod
    def _format_table(rows: list[list[str]], header: str) -> str:
        if not rows:
            return f"{header}: (empty)"

        # Limit output
        truncated = len(rows) > 100
        if truncated:
            rows = rows[:100]

        # Calculate column widths
        num_cols = max(len(r) for r in rows)
        widths = [0] * num_cols
        for row in rows:
            for j, val in enumerate(row):
                widths[j] = max(widths[j], len(str(val)[:50]))
        widths = [min(w, 50) for w in widths]

        lines = [header, ""]
        for i, row in enumerate(rows):
            cells = []
            for j in range(num_cols):
                val = str(row[j])[:50] if j < len(row) else ""
                cells.append(val.ljust(widths[j]))
            lines.append("  ".join(cells).rstrip())
            if i == 0:
                lines.append("  ".join("-" * w for w in widths))

        if truncated:
            lines.append(f"\n... (showing first 100 of {len(rows)} rows)")

        return "\n".join(lines)


class SpreadsheetWriteTool(Tool):
    """Write data to CSV or XLSX spreadsheet files."""

    def __init__(self, allowed_dir: Path | None = None):
        self._allowed_dir = allowed_dir

    @property
    def name(self) -> str:
        return "spreadsheet_write"

    @property
    def description(self) -> str:
        return (
            "Write data to a spreadsheet file (CSV or XLSX). "
            "Data should be a JSON string representing an array of arrays (rows). "
            "File type is determined by extension."
        )

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Path to the spreadsheet file (.csv or .xlsx)"},
                "sheet": {"type": "string", "description": "Sheet name (for XLSX, default: 'Sheet1')"},
                "data": {
                    "type": "string",
                    "description": 'JSON array of arrays, e.g., [["Name","Age"],["Alice",30]]',
                },
            },
            "required": ["path", "data"],
        }

    async def execute(self, **kwargs: Any) -> str:
        path_str = kwargs.get("path", "") or kwargs.get("file_path", "")
        path_str = path_str.strip() if isinstance(path_str, str) else str(path_str)
        sheet_name = kwargs.get("sheet", "")
        sheet_name = (sheet_name.strip() if isinstance(sheet_name, str) else str(sheet_name)) or "Sheet1"
        raw_data = kwargs.get("data", "")

        if not path_str:
            return "Error: No file path provided."
        if not raw_data:
            return "Error: No data provided."

        # Accept data as a list (from LLM tool calls) or JSON string
        if isinstance(raw_data, list):
            data = raw_data
        elif isinstance(raw_data, str):
            try:
                data = json.loads(raw_data.strip())
            except json.JSONDecodeError as e:
                return f"Error: Invalid JSON data: {e}"
        else:
            return "Error: Data must be a JSON array of arrays or a JSON string."

        if not isinstance(data, list) or not all(isinstance(row, list) for row in data):
            return "Error: Data must be a JSON array of arrays."

        try:
            file_path = _resolve_path(path_str, self._allowed_dir)
        except PermissionError as e:
            return f"Error: {e}"

        file_path.parent.mkdir(parents=True, exist_ok=True)
        suffix = file_path.suffix.lower()

        try:
            if suffix == ".csv":
                return self._write_csv(file_path, data)
            elif suffix in (".xlsx", ".xls"):
                return self._write_xlsx(file_path, sheet_name, data)
            else:
                return f"Error: Unsupported file type '{suffix}'. Use .csv or .xlsx."
        except Exception as e:
            return f"Error writing spreadsheet: {e}"

    @staticmethod
    def _write_csv(path: Path, data: list[list]) -> str:
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerows(data)
        return f"Wrote {len(data)} rows to {path.name}"

    @staticmethod
    def _write_xlsx(path: Path, sheet_name: str, data: list[list]) -> str:
        try:
            import openpyxl
        except ImportError:
            return "Error: openpyxl is required for XLSX files. Install with: pip install openpyxl"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        for row in data:
            ws.append(row)
        wb.save(path)
        return f"Wrote {len(data)} rows to {path.name} [{sheet_name}]"
